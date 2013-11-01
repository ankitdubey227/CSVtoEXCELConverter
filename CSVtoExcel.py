import csv
from openpyxl import Workbook
from openpyxl.cell import get_column_letter


def CSV2XLSX(file):

	f = open(file,'r')

	csv.register_dialect('comma', delimiter=',')

	reader = csv.reader(f, dialect='comma')

	wb = Workbook()
	dest_filename = file[:-3]+"xlsx"

	ws = wb.worksheets[0]
	

	for row_index, row in enumerate(reader):
		for column_index, cell in enumerate(row):
			column_letter = get_column_letter((column_index + 1))
			ws.cell('%s%s'%(column_letter, (row_index + 1))).value = cell

	wb.save(filename = dest_filename)


CSV2XLSX('sdus.csv');